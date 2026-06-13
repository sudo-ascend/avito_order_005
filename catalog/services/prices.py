from __future__ import annotations

from collections import Counter
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re

from django.core.files.base import ContentFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from catalog.models import PlantProduct, PriceUpload, SiteConfiguration

HEADERS = {
    "Сорт": "variety_name",
    "Артикул": "article",
    "ЛОКАЦИЯ": "location",
    "Тара": "container_type",
    "Кратность заказа": "order_multiple",
    "Наличие": "stock",
    "Цена": "price",
    "СКИДКА (новые)": "discount_new",
    "СКИДКА (2023-2024)": "discount_legacy",
    "СКИДКА(з+аква лого)": "discount_logo",
    "Заказ": "order_note",
}

IMPORT_FIELDS = tuple(HEADERS.values())
DECIMAL_FIELDS = {"price", "discount_new", "discount_legacy", "discount_logo"}
INTEGER_FIELDS = {"order_multiple", "stock"}


def _normalize_string(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_decimal(value):
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        return None
    return number


def _normalize_integer(value):
    if value in (None, ""):
        return None
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _normalize_value(field_name, value):
    if field_name in DECIMAL_FIELDS:
        return _normalize_decimal(value)
    if field_name in INTEGER_FIELDS:
        return _normalize_integer(value)
    return _normalize_string(value)


def _serialize_value(value):
    if isinstance(value, Decimal):
        return str(value)
    return value


def _deserialize_value(field_name, value):
    return _normalize_value(field_name, value)


def split_sort_name(raw_value: str) -> tuple[str, str]:
    raw_name = _normalize_string(raw_value)
    if not raw_name:
        return "", ""

    match = re.match(r"^(?P<latin>.+?)\s*\((?P<russian>.+)\)\s*$", raw_name)
    if not match:
        return raw_name, ""

    latin_name = _normalize_string(match.group("latin"))
    russian_name = _normalize_string(match.group("russian"))
    return russian_name or raw_name, latin_name


def compose_sort_name(russian_name: str, latin_name: str) -> str:
    russian_name = _normalize_string(russian_name)
    latin_name = _normalize_string(latin_name)
    if russian_name and latin_name:
        return f"{latin_name} ({russian_name})"
    return russian_name or latin_name


def _find_header_row(worksheet):
    for row_index in range(1, worksheet.max_row + 1):
        values = [_normalize_string(worksheet.cell(row=row_index, column=column).value) for column in range(1, worksheet.max_column + 1)]
        if "Сорт" in values and "Артикул" in values:
            return row_index, values
    raise ValueError("Не удалось найти строку заголовков в Excel файле.")


def parse_price_workbook(file_obj) -> list[dict]:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook.active
    header_row, raw_headers = _find_header_row(worksheet)
    column_map = {}
    for index, header in enumerate(raw_headers, start=1):
        if header in HEADERS:
            column_map[index] = HEADERS[header]

    parsed_rows = []
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        raw_row = {
            field_name: _normalize_value(field_name, worksheet.cell(row=row_index, column=column_index).value)
            for column_index, field_name in column_map.items()
        }
        if not raw_row.get("variety_name") and not raw_row.get("article"):
            continue
        raw_row["variety_name"], raw_row["latin_name"] = split_sort_name(raw_row.get("variety_name", ""))
        raw_row["source_row"] = row_index
        parsed_rows.append(raw_row)
    return parsed_rows


def build_change_summary(parsed_rows: list[dict]) -> dict:
    summary = Counter()
    new_records = 0

    for row in parsed_rows:
        article = row.get("article")
        if not article:
            continue
        product = PlantProduct.objects.filter(article=article).first()
        if product is None:
            new_records += 1
            for field_name in IMPORT_FIELDS:
                if row.get(field_name) not in (None, ""):
                    summary[field_name] += 1
            continue

        for field_name in IMPORT_FIELDS:
            incoming_value = row.get(field_name)
            current_value = getattr(product, field_name)
            if field_name in DECIMAL_FIELDS and current_value is not None:
                current_value = current_value.quantize(Decimal("0.01"))
            if incoming_value != current_value:
                summary[field_name] += 1

    return {
        "changes_per_field": dict(summary),
        "new_records": new_records,
        "total_rows": len(parsed_rows),
    }


def create_price_upload_from_file(*, uploaded_file, user, update_requested: bool) -> PriceUpload:
    parsed_rows = parse_price_workbook(uploaded_file)
    uploaded_file.seek(0)
    normalized_filename = PriceUpload.build_filename()
    price_upload = PriceUpload(
        original_filename=normalized_filename,
        uploaded_by=user if getattr(user, "is_authenticated", False) else None,
        update_requested=update_requested,
        row_count=len(parsed_rows),
        parsed_payload=[{key: _serialize_value(value) for key, value in row.items()} for row in parsed_rows],
        change_summary=build_change_summary(parsed_rows) if update_requested else {"total_rows": len(parsed_rows)},
    )
    price_upload.file.save(normalized_filename, ContentFile(uploaded_file.read()), save=False)
    price_upload.save()
    return price_upload


def apply_price_upload(price_upload: PriceUpload) -> int:
    applied_count = 0
    for row in price_upload.parsed_payload:
        article = _normalize_string(row.get("article"))
        if not article:
            continue

        defaults = {}
        for field_name in IMPORT_FIELDS:
            defaults[field_name] = _deserialize_value(field_name, row.get(field_name))
        defaults["latin_name"] = _normalize_string(row.get("latin_name"))
        defaults["source_row"] = row.get("source_row")

        product, _created = PlantProduct.objects.get_or_create(article=article, defaults=defaults)
        if not _created:
            for field_name, value in defaults.items():
                setattr(product, field_name, value)
            product.save()
        applied_count += 1

    price_upload.is_merged = True
    price_upload.save(update_fields=["is_merged"])
    return applied_count


def reject_price_upload(price_upload: PriceUpload):
    price_upload.is_merged = False
    price_upload.save(update_fields=["is_merged"])


def export_products_to_workbook() -> BytesIO:
    config = SiteConfiguration.objects.first() or SiteConfiguration()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"

    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    title_font = Font(name="Calibri", size=14, bold=False, color="FFFF0000")
    section_font = Font(name="Calibri", size=18, bold=True)
    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    numeric_font = Font(name="Times New Roman", size=11)

    worksheet.merge_cells("A1:C2")
    worksheet.merge_cells("A3:A4")

    worksheet["A1"] = config.price_notice
    worksheet["A3"] = config.price_catalog_title
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet["A3"].font = section_font
    worksheet["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Сорт",
        "Артикул",
        "ЛОКАЦИЯ",
        "Тара",
        "Кратность заказа",
        "Наличие",
        "Цена",
        "СКИДКА (новые)",
        "СКИДКА (2023-2024)",
        "СКИДКА(з+аква лого)",
        "Заказ",
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=7, column=column, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, product in enumerate(PlantProduct.objects.order_by("variety_name", "article"), start=8):
        values = [
            compose_sort_name(product.variety_name, product.latin_name),
            product.article,
            product.location,
            product.container_type,
            product.order_multiple,
            product.stock,
            float(product.price) if product.price is not None else None,
            float(product.discount_new) if product.discount_new is not None else None,
            float(product.discount_legacy) if product.discount_legacy is not None else None,
            float(product.discount_logo) if product.discount_logo is not None else None,
            product.order_note,
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.border = thin_border
            cell.font = numeric_font if 8 <= column <= 10 else body_font
            if column in (8, 9, 10):
                cell.number_format = "0.00"
            if column in (1, 11):
                cell.alignment = Alignment(wrap_text=True)

    last_row = max(worksheet.max_row, 7)
    worksheet.auto_filter.ref = f"A7:K{last_row}"
    worksheet.sheet_view.showGridLines = True

    worksheet.row_dimensions[1].height = 15
    worksheet.row_dimensions[2].height = 60
    worksheet.row_dimensions[4].height = 30.6
    worksheet.row_dimensions[5].height = 78
    worksheet.row_dimensions[6].height = 78
    worksheet.row_dimensions[7].height = 46.5

    widths = {"A": 73.33203125, "B": 11.6640625, "C": 13.33203125, "D": 17.33203125, "E": 17.33203125, "F": 17.33203125, "G": 5.6640625, "H": 12.33203125, "I": 11.109375, "J": 16.5546875, "K": 13.0}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
